import os
import io
import traceback
import json
import secrets
from flask import Flask, request, session, jsonify, render_template, send_file
from openpyxl import Workbook, load_workbook
from store import init_store_db, get_all_parts, add_part, update_part, delete_part, bulk_insert_parts, get_model_params, save_model_params
from auth_app import init_auth_db, register_auth_routes, require_auth, require_admin, is_admin_user

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))

register_auth_routes(app)


@app.route('/')
@require_auth
def estimator():
    return render_template('estimator.html',
                           user_name=session.get('user_name', 'User'),
                           is_admin=is_admin_user())


@app.route('/data')
@require_auth
def data_page():
    return render_template('data.html',
                           user_name=session.get('user_name', 'User'),
                           is_admin=is_admin_user())


@app.route('/api/parts')
@require_auth
def api_get_parts():
    parts = get_all_parts()
    return jsonify({'parts': parts})


@app.route('/api/parts', methods=['POST'])
@require_admin
def api_add_part():
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data'}), 400
    add_part(data, session.get('user_email', ''))
    return jsonify({'ok': True})


@app.route('/api/parts/<int:part_id>', methods=['PUT'])
@require_admin
def api_update_part(part_id):
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data'}), 400
    update_part(part_id, data)
    return jsonify({'ok': True})


@app.route('/api/parts/<int:part_id>', methods=['DELETE'])
@require_admin
def api_delete_part(part_id):
    delete_part(part_id)
    return jsonify({'ok': True})


@app.route('/api/parts/import', methods=['POST'])
@require_admin
def api_import_parts():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'File must be .xlsx'}), 400
    try:
        wb = load_workbook(f, data_only=True)
        # Try to find 'data source' sheet, fall back to active sheet
        ws = None
        for name in wb.sheetnames:
            if name.lower().strip() in ('data source', 'datasource', 'data'):
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active
        headers = [' '.join(str(c.value or '').split()).strip().lower() for c in ws[1]]
        col_map = {}
        aliases = {
            'part_number': ['part_number', 'part number', 'pn', 'part #', 'part no', 'generac part number', 'manufacturer part number'],
            'description': ['description', 'desc', 'name'],
            'process': ['process', 'mfg process', 'manufacturing process'],
            'material_family': ['material_family', 'material family', 'mat family', 'mat_family', 'general material'],
            'material': ['material', 'mat', 'material name'],
            'complexity': ['complexity', 'cx', 'complex', 'complexity (1=low, 3=high)'],
            'coo': ['coo', 'country', 'country of origin'],
            'volume_cm3': ['volume_cm3', 'volume', 'vol', 'volume (cm3)', 'vol_cm3', 'material volume (cm^3)', 'material volume (cm3)'],
            'price_hv': ['price_hv', 'price', 'hv price', 'unit price', 'cost', 'high volume pricing ($)', 'high volume pricing'],
            'tool_price': ['tool_price', 'tool price', 'tooling', 'tooling cost', 'tool cost'],
            'tool_lt': ['tool_lt', 'tool lead time', 'tool lt', 'lead time', 'tool lt (wks)'],
            'supplier': ['supplier', 'vendor', 'supply'],
            'product': ['product', 'program', 'project'],
            'revision': ['revision', 'rev'],
            'finish': ['finish', 'surface finish', 'coating'],
            'thickness_mm': ['thickness_mm', 'thickness (mm)', 'thickness', 'thk'],
            'envelope_x_mm': ['envelope_x_mm', 'envelope x (mm)', 'envelope x', 'x (mm)'],
            'envelope_y_mm': ['envelope_y_mm', 'envelope y (mm)', 'envelope y', 'y (mm)'],
            'envelope_z_mm': ['envelope_z_mm', 'envelope z (mm)', 'envelope z', 'z (mm)'],
            'production_lt': ['production_lt', 'production lt (wks)', 'production lt', 'prod lt']
        }
        for field, names in aliases.items():
            for i, h in enumerate(headers):
                if h in names:
                    col_map[field] = i
                    break
        # Detect product columns (R3, R2 INV, R2 CAB, R2 SDS, G1 ACSC, G1 MI, GB1800)
        product_cols = {}
        product_names = ['r3', 'r2 inv', 'r2 cab', 'r2 sds', 'g1 acsc', 'g1 mi', 'gb1800']
        for i, h in enumerate(headers):
            if h in product_names:
                product_cols[i] = h.upper()
        parts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            p = {}
            for field, idx in col_map.items():
                try:
                    val = row[idx] if idx < len(row) else None
                except (IndexError, TypeError):
                    val = None
                if val is None:
                    val = '' if field in ('part_number', 'description', 'process', 'material_family', 'material', 'coo', 'supplier', 'product', 'revision', 'finish') else 0
                # Ensure numeric fields are actually numeric
                if field in ('complexity', 'volume_cm3', 'price_hv', 'tool_price', 'tool_lt', 'thickness_mm', 'envelope_x_mm', 'envelope_y_mm', 'envelope_z_mm', 'production_lt'):
                    try:
                        val = float(val) if val else 0
                    except (ValueError, TypeError):
                        val = 0
                p[field] = val
            # Combine product columns into one field
            prods = []
            for idx, name in product_cols.items():
                try:
                    val = row[idx] if idx < len(row) else None
                    if val and str(val).strip().upper() in ('X', 'YES', '1', 'TRUE'):
                        prods.append(name)
                except (IndexError, TypeError):
                    pass
            if prods:
                p['product'] = ', '.join(prods)
            if p.get('process') or p.get('part_number') or p.get('description'):
                parts.append(p)
        count = bulk_insert_parts(parts, session.get('user_email', ''))
        return jsonify({'ok': True, 'imported': count})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 400


@app.route('/api/parts/export')
@require_auth
def api_export_parts():
    parts = get_all_parts()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Parts'
    headers = ['Part Number', 'Description', 'Process', 'Material Family', 'Material',
               'Complexity', 'COO', 'Volume (cm3)', 'Price HV', 'Tool Price', 'Tool LT', 'Supplier',
               'Product', 'Revision', 'Finish', 'Thickness (mm)', 'Envelope X (mm)', 'Envelope Y (mm)', 'Envelope Z (mm)', 'Production LT (wks)']
    ws.append(headers)
    for p in parts:
        ws.append([p['part_number'], p['description'], p['process'], p['material_family'],
                   p['material'], p['complexity'], p['coo'], p['volume_cm3'],
                   p['price_hv'], p['tool_price'], p['tool_lt'], p['supplier'],
                   p.get('product', ''), p.get('revision', ''), p.get('finish', ''),
                   p.get('thickness_mm', 0), p.get('envelope_x_mm', 0), p.get('envelope_y_mm', 0),
                   p.get('envelope_z_mm', 0), p.get('production_lt', 0)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='cost_estimator_parts.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/params')
@require_auth
def api_get_params():
    params = get_model_params()
    return jsonify({'params': json.loads(params) if params else None})


@app.route('/api/params', methods=['POST'])
@require_admin
def api_save_params():
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data'}), 400
    save_model_params(json.dumps(data), session.get('user_email', ''))
    return jsonify({'ok': True})


if __name__ == '__main__':
    init_auth_db()
    init_store_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
